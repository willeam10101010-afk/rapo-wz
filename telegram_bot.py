"""
Telegram Bot - Message Tracking and Validation

This bot tracks messages sent to it by saving them to an Excel file along with
user information and timestamps. It can detect duplicate messages and inform
users about when and by whom a message was previously sent.

Features:
    - Save messages to Excel file (data.xlsx)
    - Track message sender and timestamp
    - Detect duplicate messages
    - Optional GitHub integration for auto-commit

Dependencies:
    - python-telegram-bot: Telegram Bot API wrapper
    - openpyxl: Excel file handling
    - requests: HTTP client for GitHub API

Setup:
    1. Replace TOKEN with your Telegram bot token from @BotFather
    2. (Optional) Configure GitHub credentials for auto-commit feature
    3. Run: python telegram_bot.py

Author: willeam10101010-afk
License: MIT
"""

import os
import datetime
import base64
import requests
import openpyxl
from telegram import Update
from telegram.ext import ApplicationBuilder, ContextTypes, MessageHandler, filters

# ============================================
# Configuration
# ============================================

# Telegram bot token - Get this from @BotFather on Telegram
# IMPORTANT: Replace 'YOUR_BOT_TOKEN_HERE' with your actual bot token
TOKEN = 'YOUR_BOT_TOKEN_HERE'

# Path to the Excel file where messages will be stored
XLSX_FILE = 'data.xlsx'

# ============================================
# GitHub Configuration (Optional)
# ============================================
# These settings are used for the optional GitHub auto-commit feature
# If you don't need GitHub integration, you can leave these as-is

# GitHub personal access token - Create one at: https://github.com/settings/tokens
# IMPORTANT: Replace 'YOUR_GITHUB_TOKEN_HERE' with your actual token
GITHUB_TOKEN = 'YOUR_GITHUB_TOKEN_HERE'

# Repository information
REPO_OWNER = 'willeam10101010-afk'  # GitHub username or organization
REPO_NAME = 'rapo-wz'               # Repository name
BRANCH = 'main'                      # Git branch name

# ============================================
# Data Management Functions
# ============================================

def load_existing_data():
    """
    Load existing message data from the Excel file.
    
    This function reads the XLSX file and creates a dictionary mapping
    messages to their metadata (user and datetime).
    
    Returns:
        dict: A dictionary where keys are message texts and values are dicts
              containing 'user' and 'datetime' information.
              Returns an empty dict if the file doesn't exist.
    
    Example:
        {
            'Hello': {'user': 'john_doe', 'datetime': '2026-01-25 10:30:00'},
            'Test': {'user': 'jane_smith', 'datetime': '2026-01-25 11:45:00'}
        }
    """
    if os.path.exists(XLSX_FILE):
        wb = openpyxl.load_workbook(XLSX_FILE)
        sheet = wb.active
        data = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Only process rows with message text
                data[row[0]] = {'user': row[1], 'datetime': row[2]}
        return data
    return {}

def save_data(message, user, dt):
    """
    Save a new message to the Excel file.
    
    Creates a new Excel file with headers if it doesn't exist, otherwise
    appends the new message data to the existing file.
    
    Args:
        message (str): The message text to save
        user (str): The username or first name of the sender
        dt (datetime): The datetime object representing when the message was sent
    
    File Structure:
        Column A: Message (text)
        Column B: User (username/name)
        Column C: DateTime (timestamp)
    """
    if not os.path.exists(XLSX_FILE):
        # Create new workbook with headers
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet['A1'] = 'Message'
        sheet['B1'] = 'User'
        sheet['C1'] = 'DateTime'
    else:
        # Load existing workbook
        wb = openpyxl.load_workbook(XLSX_FILE)
        sheet = wb.active
    
    # Append new data row
    sheet.append([message, user, str(dt)])
    wb.save(XLSX_FILE)

# ============================================
# GitHub Integration (Optional)
# ============================================

def commit_file_to_github(file_path, commit_message):
    """
    Commit a file to the GitHub repository.
    
    This function uploads a file to GitHub using the GitHub API. It handles
    both creating new files and updating existing files.
    
    Args:
        file_path (str): Path to the local file to commit
        commit_message (str): Commit message for the GitHub commit
    
    Returns:
        str: Success or error message describing the result
    
    Note:
        - Requires valid GITHUB_TOKEN, REPO_OWNER, and REPO_NAME configuration
        - The file will be committed to the branch specified in BRANCH variable
        - If the file already exists in the repository, it will be updated
    
    Example:
        result = commit_file_to_github('data.xlsx', 'Update message data')
        print(result)  # "File berhasil di-commit ke repo."
    """
    try:
        # Read file content and encode to base64
        with open(file_path, 'rb') as f:
            content = base64.b64encode(f.read()).decode('utf-8')
        
        # Prepare GitHub API request
        url = f'https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/contents/{file_path}'
        headers = {'Authorization': f'token {GITHUB_TOKEN}'}
        
        # Check if file already exists to get its SHA
        response = requests.get(url, headers=headers)
        data = {
            'message': commit_message,
            'content': content,
            'branch': BRANCH
        }
        
        # If file exists, include its SHA for update operation
        if response.status_code == 200:
            data['sha'] = response.json()['sha']
        
        # Commit the file (create or update)
        response = requests.put(url, headers=headers, json=data)
        if response.status_code in [200, 201]:
            return "File berhasil di-commit ke repo."
        else:
            return f"Gagal commit: {response.json()}"
    except Exception as e:
        return f"Error: {str(e)}"

# ============================================
# Message Handler
# ============================================

async def message_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Handle incoming text messages from users.
    
    This is the main message processing function that:
    1. Extracts message text and user information
    2. Checks if the message has been sent before
    3. Responds appropriately based on whether it's a duplicate
    4. Saves new messages to the Excel file
    
    Args:
        update (Update): Telegram update object containing the message
        context (ContextTypes.DEFAULT_TYPE): Context object for the handler
    
    Behavior:
        - For duplicate messages: Replies with when and by whom it was sent before
        - For new messages: Saves to Excel and confirms the data can be used
    
    Message Responses (in Indonesian):
        - Duplicate: "Pesan data pernah dikirim oleh {user} pada {datetime}"
        - New: "Data dapat digunakan."
    """
    # Extract message details
    user_message = update.message.text
    user = update.message.from_user.username or update.message.from_user.first_name
    dt = datetime.datetime.now()
    
    # Load existing messages from Excel
    existing_data = load_existing_data()
    
    # Check if message is a duplicate
    if user_message in existing_data:
        info = existing_data[user_message]
        reply = f"Pesan data pernah dikirim oleh {info['user']} pada {info['datetime']}"
        await update.message.reply_text(reply)
    else:
        # New message - save it
        await update.message.reply_text("Data dapat digunakan.")
        save_data(user_message, user, dt)
        
        # Optional: Uncomment the line below to enable GitHub auto-commit
        # commit_file_to_github(XLSX_FILE, f'New message from {user}')

# ============================================
# Application Setup and Entry Point
# ============================================

def main():
    """
    Initialize and start the Telegram bot.
    
    This function sets up the bot application, registers the message handler,
    and starts polling for messages from Telegram.
    
    The bot will:
    1. Connect to Telegram using the provided TOKEN
    2. Listen for text messages
    3. Process messages using the message_handler function
    4. Continue running until manually stopped (Ctrl+C)
    """
    # Build the application with the bot token
    application = ApplicationBuilder().token(TOKEN).build()
    
    # Register message handler for text messages
    application.add_handler(MessageHandler(filters.TEXT, message_handler))
    
    # Start polling for messages
    print("Bot started successfully! Waiting for messages...")
    print("Press Ctrl+C to stop the bot.")
    application.run_polling()

if __name__ == '__main__':
    main()
